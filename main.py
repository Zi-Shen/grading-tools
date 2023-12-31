import subprocess
import os
import openpyxl

# path of the folder that contains all the assignments
path = "D:/Desktop/project/grading-tools/COMP2121_1B_2023-Assignment-4"

# file names as a list in the above path
filelist = os.listdir(path)
counter = 0

grade_book_file = 'COMP2121_1B_2023 Assignment_4 Grades.xlsx'
workbook = openpyxl.load_workbook(grade_book_file)

sheet = workbook.active

def find_name_in_grade_book(name):
    last_name = name.split(' ')[0]
    first_name = ' '.join(name.split(' ')[1:])
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 2).value == last_name and sheet.cell(row, 1).value == first_name:
            return row
    print(f'Cannot find {name} in the grade book.')
    return None

counter_graded = 0
for name_info in filelist:
    temp_path = path + "/" + name_info + "/"
    path_pdf = temp_path + os.listdir(temp_path)[0]
    path_pdf = path_pdf.replace(' ', '%20')
    name = name_info.split('_')[0]
    
    row = find_name_in_grade_book(name)
    counter_graded += 1

    if sheet.cell(row, 16).value is not None or sheet.cell(row, 9).value is not None or sheet.cell(row, 13).value == 0:
        if sheet.cell(row, 13).value == 0 or sheet.cell(row, 13).value == '0':
            print(f'{name} has not submitted the assignment before 2 day after deadline. Skip grading...')
            continue
        continue

    print(f'{counter_graded}/{len(filelist)} has been graded, good job!')
    print(f'Start grading {name}...')

    subprocess.run(f'start msedge file:///{path_pdf}', shell=True)

    # input the grade and feedback
    feedback = input("Input the feedback:\n")
    grade = input("Input the grade:\n")

    # write the grade and feedback to the file

    sheet.cell(row, 9).value = float(grade)
    sheet.cell(row, 16).value = feedback
    workbook.save(grade_book_file)
    counter += 1
    print(f'{name} has been graded.\n')

    # kill the msedge process every 6 assignments
    if counter % 6 == 0:
        subprocess.run('taskkill /f /im msedge.exe', shell=True)
